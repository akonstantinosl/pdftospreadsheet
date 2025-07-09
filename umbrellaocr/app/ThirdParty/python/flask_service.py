from flask import Flask, request, jsonify
from flask_cors import CORS
import base64
import tempfile
import os
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np
from PIL import Image
import pdf2image
import cv2
from collections import defaultdict
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from rapidocr_onnxruntime import RapidOCR

app = Flask(__name__)
CORS(app)

# Define the path to the shared folder on the server
SHARED_FOLDER = '/srv/samba/converter_files' 

# Initialize OCR engine
engine = RapidOCR(
    rec_batch_num=6,
    det_model_name='ch_PP-OCRv4_det',
    cls_model_name='ch_ppocr_mobile_v2.0_cls',
    rec_model_name='ch_PP-OCRv4_rec',
    use_angle_cls=True,
    det_limit_side_len=960,
    det_db_thresh=0.3,
    det_db_box_thresh=0.5,
    det_db_unclip_ratio=1.6
)

def preprocess_image_for_ocr(img_array):
    """Preprocess image to improve OCR results for small text."""
    if len(img_array.shape) == 3:
        gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
    else:
        gray = img_array.copy()
    
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
        cv2.THRESH_BINARY, 11, 2
    )
    
    kernel = np.ones((2, 2), np.uint8)
    dilated = cv2.dilate(binary, kernel, iterations=1)
    
    edge_enhanced = cv2.Laplacian(gray, cv2.CV_8U, ksize=3)
    sharpened = cv2.addWeighted(gray, 1.5, edge_enhanced, -0.5, 0)
    
    return sharpened

def detect_lines(img_gray, min_length=50, orientation='horizontal', line_type='any'):
    """Detects lines in an image with enhanced support for dotted and dashed lines."""
    height, width = img_gray.shape
    
    binary = cv2.adaptiveThreshold(
        img_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
        cv2.THRESH_BINARY_INV, 11, 2
    )
    
    if line_type in ['dotted', 'any']:
        if orientation == 'horizontal':
            profile = np.sum(binary, axis=1)
            smooth_profile = gaussian_filter1d(profile, sigma=2)
            lines_img = np.zeros_like(binary)
            peaks, _ = find_peaks(smooth_profile, height=width*0.1, distance=15)
            
            for peak in peaks:
                if peak > 0 and peak < height:
                    lines_img[peak, :] = 255
        else:
            profile = np.sum(binary, axis=0)
            smooth_profile = gaussian_filter1d(profile, sigma=2)
            lines_img = np.zeros_like(binary)
            peaks, _ = find_peaks(smooth_profile, height=height*0.1, distance=15)
            
            for peak in peaks:
                if peak > 0 and peak < width:
                    lines_img[:, peak] = 255
        
        if line_type == 'any':
            if orientation == 'horizontal':
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_length, 1))
            else:
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, min_length))
            
            morph_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
            combined_lines = cv2.bitwise_or(lines_img, morph_lines)
            return combined_lines
        
        return lines_img
    else:
        if orientation == 'horizontal':
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_length, 1))
        else:
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, min_length))
        
        lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
        return lines

def detect_tables(img_array, line_sensitivity=25, use_grid=True, dotted_line_support=True):
    """Detects tables in an image with enhanced support for dotted/dashed lines."""
    gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
    
    if use_grid:
        if dotted_line_support:
            horizontal_lines = detect_lines(gray, min_length=line_sensitivity, orientation='horizontal', line_type='any')
            vertical_lines = detect_lines(gray, min_length=line_sensitivity, orientation='vertical', line_type='any')
        else:
            binary = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                cv2.THRESH_BINARY_INV, 11, 2
            )
            
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (line_sensitivity, 1))
            horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, line_sensitivity))
            vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        
        table_grid = cv2.bitwise_or(horizontal_lines, vertical_lines)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        dilated_grid = cv2.dilate(table_grid, kernel, iterations=2)
        
        contours, _ = cv2.findContours(dilated_grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        table_contours = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            area = w * h
            if area > (img_array.shape[0] * img_array.shape[1]) / 30:
                aspect_ratio = w / float(h)
                if 0.2 < aspect_ratio < 5:
                    table_contours.append((x, y, w, h, horizontal_lines, vertical_lines))
        
        if table_contours:
            return table_contours
    
    return text_based_table_detection(img_array, gray)

def text_based_table_detection(img_array, gray):
    """Detects tables based on text alignment and clustering when line detection fails."""
    preprocessed_img = preprocess_image_for_ocr(img_array)
    ocr_result = engine(preprocessed_img)
    
    if not ocr_result or len(ocr_result[0]) == 0:
        return []
    
    text_boxes = []
    for box, text, confidence in ocr_result[0]:
        if text.strip():
            x_coords = [p[0] for p in box]
            y_coords = [p[1] for p in box]
            x_min, y_min = min(x_coords), min(y_coords)
            x_max, y_max = max(x_coords), max(y_coords)
            width, height = x_max - x_min, y_max - y_min
            
            text_boxes.append({
                'box': box,
                'text': text,
                'confidence': confidence,
                'x': x_min,
                'y': y_min,
                'width': width,
                'height': height,
                'x_center': (x_min + x_max) / 2,
                'y_center': (y_min + y_max) / 2
            })
    
    if not text_boxes:
        return []
    
    text_boxes_sorted_y = sorted(text_boxes, key=lambda b: b['y'])
    avg_height = sum(box['height'] for box in text_boxes) / len(text_boxes)
    row_threshold = max(avg_height * 1.5, 15)
    
    rows = []
    current_row = [text_boxes_sorted_y[0]]
    current_y = text_boxes_sorted_y[0]['y_center']
    
    for box in text_boxes_sorted_y[1:]:
        if abs(box['y_center'] - current_y) <= row_threshold:
            current_row.append(box)
        else:
            rows.append(current_row)
            current_row = [box]
            current_y = box['y_center']
    
    if current_row:
        rows.append(current_row)
    
    if len(rows) < 2:
        return []
    
    table_candidates = []
    min_rows_for_table = 3
    for i in range(len(rows) - min_rows_for_table + 1):
        for k in range(len(rows) - i - min_rows_for_table + 1):
            potential_table_rows = rows[i:i+min_rows_for_table+k]
            row_lengths = [len(row) for row in potential_table_rows]
            
            if max(row_lengths) - min(row_lengths) <= 1:
                all_boxes = [box for row in potential_table_rows for box in row]
                x_min = min(box['x'] for box in all_boxes)
                y_min = min(box['y'] for box in all_boxes)
                x_max = max(box['x'] + box['width'] for box in all_boxes)
                y_max = max(box['y'] + box['height'] for box in all_boxes)
                
                x_min = max(0, x_min - 10)
                y_min = max(0, y_min - 10)
                x_max = min(img_array.shape[1], x_max + 10)
                y_max = min(img_array.shape[0], y_max + 10)
                
                width = x_max - x_min
                height = y_max - y_min
                
                empty_lines = np.zeros_like(gray)
                table_candidates.append((int(x_min), int(y_min), int(width), int(height), empty_lines, empty_lines))
    
    return table_candidates

def detect_cells_from_text(text_items, table_width, table_height):
    """Detects cell structure from text positions when explicit grid lines are missing."""
    if not text_items:
        return None, None
    
    avg_height = sum(item['height'] for item in text_items) / len(text_items)
    avg_width = sum(item['width'] for item in text_items) / len(text_items)
    
    sorted_by_y = sorted(text_items, key=lambda t: t['y_center'])
    y_centers = [t['y_center'] for t in sorted_by_y]
    y_threshold = max(avg_height * 0.8, 10)
    
    def cluster_coordinates(coords, threshold):
        clusters = []
        current_cluster = [coords[0]]
        current_value = coords[0]
        
        for val in coords[1:]:
            if abs(val - current_value) <= threshold:
                current_cluster.append(val)
            else:
                if current_cluster:
                    clusters.append(sum(current_cluster) / len(current_cluster))
                current_cluster = [val]
                current_value = val
        
        if current_cluster:
            clusters.append(sum(current_cluster) / len(current_cluster))
        
        return clusters
    
    row_centers = cluster_coordinates(y_centers, y_threshold)
    row_boundaries = [0]
    for i in range(len(row_centers) - 1):
        boundary = int((row_centers[i] + row_centers[i+1]) / 2)
        row_boundaries.append(boundary)
    row_boundaries.append(int(table_height))
    
    sorted_by_x = sorted(text_items, key=lambda t: t['x_center'])
    x_centers = [t['x_center'] for t in sorted_by_x]
    x_threshold = max(avg_width * 0.8, 20)
    
    col_centers = cluster_coordinates(x_centers, x_threshold)
    col_boundaries = [0]
    for i in range(len(col_centers) - 1):
        boundary = int((col_centers[i] + col_centers[i+1]) / 2)
        col_boundaries.append(boundary)
    col_boundaries.append(int(table_width))
    
    return row_boundaries, col_boundaries

def process_table_ocr(ocr_results, table_rect, use_inferred_grid=True):
    """Process OCR results to reconstruct a table."""
    x_table, y_table, w_table, h_table = table_rect[:4]
    
    if not ocr_results or len(ocr_results[0]) == 0:
        return None
    
    table_texts = []
    for box, text, confidence in ocr_results[0]:
        x_center = sum([p[0] for p in box]) / 4
        y_center = sum([p[1] for p in box]) / 4
        
        if ((x_table - 5 <= x_center <= x_table + w_table + 5) and 
            (y_table - 5 <= y_center <= y_table + h_table + 5)):
            x_min = min([p[0] for p in box])
            y_min = min([p[1] for p in box])
            width = max([p[0] for p in box]) - x_min
            height = max([p[1] for p in box]) - y_min
            
            x_rel = x_min - x_table
            y_rel = y_min - y_table
            
            if text.strip():
                table_texts.append({
                    'text': text.strip(),
                    'confidence': confidence,
                    'x': x_rel,
                    'y': y_rel,
                    'width': width,
                    'height': height,
                    'x_center': x_center - x_table,
                    'y_center': y_center - y_table,
                    'original_box': box
                })
    
    if not table_texts:
        return None
    
    if use_inferred_grid:
        row_boundaries, col_boundaries = detect_cells_from_text(table_texts, w_table, h_table)
        
        if row_boundaries and col_boundaries and len(row_boundaries) > 1 and len(col_boundaries) > 1:
            num_rows = len(row_boundaries) - 1
            num_cols = len(col_boundaries) - 1
            table_data = [[''] * num_cols for _ in range(num_rows)]
            
            for text_item in table_texts:
                row_idx = None
                for i in range(len(row_boundaries) - 1):
                    if row_boundaries[i] <= text_item['y_center'] < row_boundaries[i+1]:
                        row_idx = i
                        break
                
                col_idx = None
                for i in range(len(col_boundaries) - 1):
                    if col_boundaries[i] <= text_item['x_center'] < col_boundaries[i+1]:
                        col_idx = i
                        break
                
                if row_idx is not None and col_idx is not None:
                    if row_idx < num_rows and col_idx < num_cols:
                        if table_data[row_idx][col_idx]:
                            table_data[row_idx][col_idx] += ' ' + text_item['text']
                        else:
                            table_data[row_idx][col_idx] = text_item['text']
            
            return table_data
    
    return reconstruct_table(table_texts, w_table, h_table)

def reconstruct_table(texts, table_width, table_height):
    """Reconstructs table structure from text positions."""
    if not texts:
        return None
    
    texts_sorted_by_y = sorted(texts, key=lambda t: t['y'])
    avg_text_height = sum(t['height'] for t in texts) / len(texts)
    row_threshold = max(avg_text_height * 0.8, 12)
    
    rows = []
    current_row = [texts_sorted_by_y[0]]
    row_y = texts_sorted_by_y[0]['y_center']
    
    for t in texts_sorted_by_y[1:]:
        if abs(t['y_center'] - row_y) <= row_threshold:
            current_row.append(t)
        else:
            if current_row:
                current_row = sorted(current_row, key=lambda t: t['x'])
                rows.append(current_row)
            current_row = [t]
            row_y = t['y_center']
    
    if current_row:
        current_row = sorted(current_row, key=lambda t: t['x'])
        rows.append(current_row)
    
    all_x_centers = []
    for row in rows:
        for t in row:
            all_x_centers.append(t['x_center'])
    
    def cluster_x_centers(x_centers):
        if not x_centers:
            return []
        
        avg_text_width = sum(t['width'] for t in texts) / len(texts)
        col_threshold = max(avg_text_width * 1.5, 20)
        
        x_centers = sorted(x_centers)
        clusters = [[x_centers[0]]]
        
        for x in x_centers[1:]:
            if abs(x - clusters[-1][-1]) <= col_threshold:
                clusters[-1].append(x)
            else:
                clusters.append([x])
        
        return [sum(cluster) / len(cluster) for cluster in clusters]
    
    col_centers = cluster_x_centers(all_x_centers)
    num_cols = len(col_centers)
    
    if num_cols == 0:
        num_cols = 1
        col_centers = [table_width / 2]
    
    table_data = []
    
    for row_texts in rows:
        row_data = [''] * num_cols
        
        for text in row_texts:
            col_idx = min(range(num_cols), 
                       key=lambda i: abs(text['x_center'] - col_centers[i]))
            
            if row_data[col_idx]:
                row_data[col_idx] += ' ' + text['text']
            else:
                row_data[col_idx] = text['text']
        
        table_data.append(row_data)
    
    return table_data

def create_tables_export(tables, output_format='xlsx'):
    """Creates a file from extracted tables in the specified format."""
    output = io.BytesIO()
    
    def calculate_column_widths(dataframe):
        widths = []
        for col in dataframe.columns:
            max_length = 0
            for entry in dataframe[col].astype(str):
                max_length = max(max_length, len(entry))
            column_width = max(max_length, len(str(col)))
            widths.append(column_width + 2)
        return widths
    
    if output_format == 'xlsx':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for i, df in enumerate(tables):
                sheet_name = f"Table {i+1}"
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(df.columns):
                    max_length = 0
                    for entry in df[col].astype(str):
                        max_length = max(max_length, len(entry))
                    column_width = max(max_length, len(str(col)))
                    worksheet.column_dimensions[get_column_letter(idx+1)].width = column_width + 2
    
    elif output_format == 'csv':
        if len(tables) == 1:
            tables[0].to_csv(output, index=False, header=False)
        else:
            import zipfile
            with zipfile.ZipFile(output, 'w') as zf:
                for i, df in enumerate(tables):
                    file_name = f"table_{i+1}.csv"
                    temp_csv = io.StringIO()
                    df.to_csv(temp_csv, index=False, header=False)
                    zf.writestr(file_name, temp_csv.getvalue())
    
    elif output_format == 'ods':
        # Enhanced ODS implementation with proper column width support
        try:
            # Method 1: Try using pandas with odfpy engine
            with pd.ExcelWriter(output, engine='odf') as writer:
                for i, df in enumerate(tables):
                    sheet_name = f"Table {i+1}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
        except ImportError:
            # Method 2: Use odfpy directly for better control over formatting
            try:
                from odf.opendocument import OpenDocumentSpreadsheet
                from odf.style import Style, TableColumnProperties, TableCellProperties
                from odf.table import Table, TableColumn, TableRow, TableCell
                from odf.text import P
                from odf import teletype
                
                # Create new ODS document
                doc = OpenDocumentSpreadsheet()
                
                for i, df in enumerate(tables):
                    sheet_name = f"Table{i+1}"  # Remove space for ODS compatibility
                    
                    # Create table
                    table = Table(name=sheet_name)
                    doc.spreadsheet.addElement(table)
                    
                    # Calculate column widths based on content
                    col_widths = []
                    for col_idx, col_name in enumerate(df.columns):
                        max_length = 0
                        # Check column data for max length
                        for _, row in df.iterrows():
                            cell_value = str(row.iloc[col_idx]) if pd.notna(row.iloc[col_idx]) else ""
                            max_length = max(max_length, len(cell_value))
                        
                        # Set minimum width and add padding
                        col_width = max(max_length, 8) + 2  # minimum 8 chars + 2 padding
                        col_widths.append(col_width)
                    
                    # Create column styles with proper widths
                    for col_idx, width in enumerate(col_widths):
                        # Create style for this column
                        style_name = f"col{i}_{col_idx}_style"
                        col_style = Style(name=style_name, family="table-column")
                        
                        # Convert character width to cm (approx. 0.25cm per char for better readability)
                        width_cm = width * 0.25
                        col_style.addElement(TableColumnProperties(columnwidth=f"{width_cm:.2f}cm"))
                        doc.automaticstyles.addElement(col_style)
                        
                        # Add column with style
                        table.addElement(TableColumn(stylename=style_name))
                    
                    # Add data rows
                    for _, row in df.iterrows():
                        tr = TableRow()
                        table.addElement(tr)
                        
                        for cell_value in row:
                            tc = TableCell()
                            tr.addElement(tc)
                            
                            # Handle different value types and ensure proper text formatting
                            if pd.notna(cell_value):
                                cell_text = str(cell_value).strip()
                                if cell_text:  # Only add non-empty text
                                    p = P()
                                    teletype.addTextToElement(p, cell_text)
                                    tc.addElement(p)
                                else:
                                    # Add empty paragraph for empty cells
                                    tc.addElement(P())
                            else:
                                # Add empty paragraph for null cells
                                tc.addElement(P())
                
                # Save to output buffer
                doc.save(output)
                
            except ImportError:
                # Method 3: Fallback to pyexcel-ods3
                try:
                    import pyexcel_ods3
                    
                    # Prepare data structure for ODS export
                    ods_data = {}
                    for i, df in enumerate(tables):
                        sheet_name = f"Table {i+1}"
                        # Convert DataFrame to list of lists, ensuring all values are properly formatted
                        sheet_data = []
                        for _, row in df.iterrows():
                            row_data = []
                            for cell in row:
                                if pd.notna(cell):
                                    # Clean and format cell data
                                    cell_str = str(cell).strip()
                                    row_data.append(cell_str)
                                else:
                                    row_data.append('')
                            sheet_data.append(row_data)
                        ods_data[sheet_name] = sheet_data
                    
                    # Write to output buffer
                    pyexcel_ods3.save_data(output, ods_data)
                    
                except ImportError:
                    # Final fallback - create Excel format
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for i, df in enumerate(tables):
                            sheet_name = f"Table {i+1}"
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                            
                            worksheet = writer.sheets[sheet_name]
                            for idx, col in enumerate(df.columns):
                                max_length = 0
                                for entry in df[col].astype(str):
                                    max_length = max(max_length, len(entry))
                                column_width = max(max_length, len(str(col)))
                                worksheet.column_dimensions[get_column_letter(idx+1)].width = column_width + 2
    
    return output.getvalue()

def process_image_for_tables(img_array):
    """Process an image to detect and extract tables."""
    preprocessed_img = preprocess_image_for_ocr(img_array)
    ocr_result = engine(preprocessed_img)
    
    table_rects = detect_tables(img_array, 25, True, True)
    
    if not table_rects:
        if isinstance(ocr_result, tuple) and len(ocr_result) >= 1 and ocr_result[0]:
            all_texts = [item[1] for item in ocr_result[0]]
            return [pd.DataFrame({"Text": all_texts})]
        return []
    
    tables = []
    for table_rect in table_rects:
        table_data = process_table_ocr(ocr_result, table_rect, use_inferred_grid=True)
        
        if table_data:
            df = pd.DataFrame(table_data)
            tables.append(df)
    
    return tables

def process_pdf(file_content):
    """Process a PDF file and extract tables from all pages."""
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
        tmp_file.write(file_content)
        tmp_file.flush()
        
        try:
            pdf_images = pdf2image.convert_from_path(tmp_file.name, dpi=300)
            
            all_tables = []
            for i, img in enumerate(pdf_images):
                img_array = np.array(img.convert('RGB'))
                page_tables = process_image_for_tables(img_array)
                all_tables.extend(page_tables)
            
            return all_tables
        finally:
            os.unlink(tmp_file.name)

def process_image(file_content):
    """Process an image file and extract tables."""
    with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
        tmp_file.write(file_content)
        tmp_file.flush()
        
        try:
            image = Image.open(tmp_file.name)
            img_array = np.array(image.convert('RGB'))
            tables = process_image_for_tables(img_array)
            return tables
        finally:
            os.unlink(tmp_file.name)

@app.route('/process', methods=['POST'])
def process_file():
    try:
        data = request.get_json()
        
        if not data or 'filename' not in data:
            return jsonify({
                'success': False,
                'message': 'No filename provided'
            }), 400

        filename = data.get('filename')
        output_format = data.get('format', 'xlsx')
        
        # Construct the full path to the input file in the shared folder
        input_filepath = os.path.join(SHARED_FOLDER, filename)

        if not os.path.exists(input_filepath):
            return jsonify({
                'success': False,
                'message': f'File not found on server: {filename}'
            }), 404
        
        # Read the file content directly from the path
        with open(input_filepath, 'rb') as f:
            file_content = f.read()

        file_extension = filename.lower().split('.')[-1] if '.' in filename else ''
        
        if file_extension == 'pdf':
            tables = process_pdf(file_content)
        elif file_extension in ['jpg', 'jpeg', 'png']:
            tables = process_image(file_content)
        else:
            return jsonify({
                'success': False,
                'message': 'Unsupported file type'
            }), 400
        
        # Create export file data in memory
        export_data = create_tables_export(tables, output_format)
        
        # Define the output filename
        original_basename = os.path.splitext(filename)[0]
        output_extension = output_format
        if output_format == 'csv' and len(tables) > 1:
            output_extension = 'zip'
        
        # A unique ID is still good to avoid filename collisions for the output
        output_filename = f"processed_{original_basename}.{output_extension}"
        output_filepath = os.path.join(SHARED_FOLDER, output_filename)
        
        # Write the processed file to the shared folder
        with open(output_filepath, 'wb') as f:
            f.write(export_data)

        # Clean up the original uploaded file
        os.remove(input_filepath)
        
        return jsonify({
            'success': True,
            'message': 'Conversion completed successfully',
            'output_filename': output_filename, # Return the filename of the result
            'tables_count': len(tables)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Processing error: {str(e)}'
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'message': 'Flask service is running'
    })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)